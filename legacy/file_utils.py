"""
file_utils.py — File I/O utilities.

Functions:
  read_sub           — Read SRT file content as string
  get_pairs          — Pair videos with SRT files by filename stem
  read_local_input   — Read visible rows from Local Import Excel file (openpyxl)
  filter_sub_excluding_approved — Insert [ALREADY USED] tags into SRT
  create_clip_subtitle          — Extract SRT lines for a clip's time range
  validate_timestamp — Validate HH:MM:SS format + logic
  expand_sub_context — Expand start/end by ±N dialogue lines from SRT
  save_log           — Save log.json
  load_log           — Load log.json
"""

import os
import re
import json
from datetime import datetime

import openpyxl


# ---------------------------------------------------------------------------
# SRT helpers
# ---------------------------------------------------------------------------

def read_sub(file_path: str) -> str:
    """Read SRT file and return content as string."""
    with open(file_path, "r", encoding="utf-8") as f:
        return f.read()


def get_pairs(video_dir: str, sub_dir: str) -> list[dict]:
    """
    Pair videos and SRT files by filename stem.
    Supports .mp4, .mkv, .avi, .mov video files and .srt sub files.
    Skips pairs where SRT is missing — logs warning, does not crash.

    Returns list of {"video": abs_path, "srt": abs_path}
    """
    supported_video_ext = {".mp4", ".mkv", ".avi", ".mov"}
    pairs = []

    for filename in sorted(os.listdir(video_dir)):
        ext = os.path.splitext(filename)[1].lower()
        if ext not in supported_video_ext:
            continue

        stem = os.path.splitext(filename)[0]
        srt_path = os.path.join(sub_dir, stem + ".srt")

        if not os.path.exists(srt_path):
            print(f"[file_utils] Warning: No SRT found for {filename} — skipping")
            continue

        pairs.append({
            "video": os.path.join(video_dir, filename),
            "srt": srt_path,
        })

    return pairs


# ---------------------------------------------------------------------------
# Local Import
# ---------------------------------------------------------------------------

def read_local_input(excel_path: str) -> list[dict]:
    """
    Read visible rows from Local Import Excel file using openpyxl.
    Skips hidden rows (row.hidden == True).
    Skips rows with invalid timestamp format — logs warning, does not crash.

    Expected columns (case-insensitive): video, start, end, language_item, language_tag, original_word
    Returns list of dicts with ai_note="" and example_sub="" added.
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Build header map from first row
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[str(cell.value).strip().lower()] = col_idx

    required = {"video", "start", "end"}
    missing = required - set(headers.keys())
    if missing:
        raise ValueError(f"Local Import file missing required columns: {missing}")

    results = []
    for row in ws.iter_rows(min_row=2):
        # Skip hidden rows
        if ws.row_dimensions[row[0].row].hidden:
            continue
        # Skip fully empty rows
        values = [cell.value for cell in row]
        if all(v is None for v in values):
            continue

        def get_col(col_name: str) -> str:
            idx = headers.get(col_name)
            if idx is None:
                return ""
            val = row[idx - 1].value
            return str(val).strip() if val is not None else ""

        def get_raw(col_name: str):
            idx = headers.get(col_name)
            return row[idx - 1].value if idx is not None else None

        video = get_col("video")
        # Normalize raw cell (datetime.time / short '0:5:32') → 'HH:MM:SS' so a
        # correctly-meant-but-loosely-typed time is accepted instead of skipped.
        start = normalize_timestamp(get_raw("start"))
        end = normalize_timestamp(get_raw("end"))

        if not video or not start or not end:
            print(f"[file_utils] Warning: Skipping row with missing video/start/end")
            continue

        if not _is_valid_timestamp(start) or not _is_valid_timestamp(end):
            print(f"[file_utils] Warning: Invalid timestamp in row ({start} / {end}) — skipping")
            continue

        results.append({
            "video": video,
            "start": start,
            "end": end,
            "language_item": get_col("language_item"),
            "language_tag": get_col("language_tag"),
            "original_word": get_col("original_word"),
            "ai_note": "",
            "example_sub": "",
        })

    return results


# ---------------------------------------------------------------------------
# SRT manipulation
# ---------------------------------------------------------------------------

def filter_sub_excluding_approved(srt_content: str, approved_segments: list[dict]) -> str:
    """
    Insert [ALREADY USED] tags into SRT before blocks matching approved start timestamps.

    approved_segments: list of {"start": "HH:MM:SS", "end": "HH:MM:SS", "language_item": str}
    Match by start timestamp (HH:MM:SS, ignoring milliseconds).

    Returns modified SRT string. Original content preserved — no lines deleted.
    """
    # Build lookup: start_hms → segment info
    used_map = {}
    for seg in approved_segments:
        start_hms = seg.get("start", "").strip()
        if start_hms:
            end_hms = seg.get("end", "").strip()
            language_item = seg.get("language_item", "").strip()
            used_map[start_hms] = {"end": end_hms, "language_item": language_item}

    if not used_map:
        return srt_content

    lines = srt_content.splitlines(keepends=True)
    output = []
    i = 0
    # SRT timestamp pattern: 00:00:00,000 --> 00:00:00,000
    ts_pattern = re.compile(
        r"(\d{2}:\d{2}:\d{2}),\d{3}\s*-->\s*(\d{2}:\d{2}:\d{2}),\d{3}"
    )

    while i < len(lines):
        line = lines[i]
        match = ts_pattern.search(line)
        if match:
            start_hms = match.group(1)
            if start_hms in used_map:
                seg_info = used_map[start_hms]
                language_item = seg_info["language_item"]
                tag = (
                    f'[ALREADY USED: from {start_hms} to {seg_info["end"]} '
                    f'| phrases used within this frame: "{language_item}"]\n'
                )
                output.append(tag)
        output.append(line)
        i += 1

    return "".join(output)


def _norm_match_local(s: str) -> str:
    """Lowercase + drop punctuation/apostrophes, collapse whitespace.
    Mirrors output_utils._norm_match so highlight & [ALREADY USED] matching agree."""
    return " ".join(re.findall(r"[a-z0-9]+", str(s).lower()))


def _block_text(raw: str) -> str:
    """Extract just the dialogue lines from an SRT block's raw text
    (drop the index line and the timestamp line)."""
    out = []
    seen_ts = False
    for ln in raw.splitlines():
        if "-->" in ln:
            seen_ts = True
            continue
        if seen_ts:
            out.append(ln)
    return " ".join(out)


def slice_srt_by_time(srt_content: str, start_hms: str, end_hms: str) -> str:
    """
    Return SRT text of blocks overlapping [start_hms, end_hms], KEEPING the original
    timestamps (unlike create_clip_subtitle which resets to 0-based).
    Used by scoped Find More to send only a time window to the AI.
    Robust: blank range or any error → return original content unchanged.
    """
    try:
        if not start_hms or not end_hms:
            return srt_content
        start_sec = _hms_to_seconds(start_hms)
        end_sec = _hms_to_seconds(end_hms)
        blocks = _parse_srt_blocks(srt_content)
        if not blocks:
            return srt_content
        kept = [b["raw"].strip() for b in blocks
                if b["end_sec"] > start_sec and b["start_sec"] < end_sec]
        return ("\n\n".join(kept) + "\n") if kept else srt_content
    except Exception:
        return srt_content


def filter_sub_excluding_approved_by_item(srt_content: str, approved_segments: list[dict]) -> str:
    """
    Like filter_sub_excluding_approved, but anchor the [ALREADY USED] tag to the
    subtitle block that actually CONTAINS the language_item (within the segment range).
    Falls back to the segment's start block. Used for scoped Find More so used content
    is still flagged even when the segment's start lies outside the trimmed window.
    Robust: any error → degrade to filter_sub_excluding_approved → original content.
    """
    try:
        blocks = _parse_srt_blocks(srt_content)
        if not blocks or not approved_segments:
            return srt_content
        tag_for_block = {}   # block index -> tag string
        used_idx = set()
        for seg in approved_segments:
            start_hms = (seg.get("start") or "").strip()
            end_hms = (seg.get("end") or "").strip()
            language_item = (seg.get("language_item") or "").strip()
            if not start_hms:
                continue
            try:
                s_sec = _hms_to_seconds(start_hms)
                e_sec = _hms_to_seconds(end_hms) if end_hms else s_sec
            except Exception:
                continue
            focus = _norm_match_local(language_item)
            chosen = None
            # 1) block within [start,end] whose dialogue contains the language_item
            if focus:
                for bi, b in enumerate(blocks):
                    if bi in used_idx:
                        continue
                    if b["end_sec"] > s_sec and b["start_sec"] < e_sec + 0.001:
                        if focus in _norm_match_local(_block_text(b["raw"])):
                            chosen = bi
                            break
            # 2) fallback: block whose start matches the segment start
            if chosen is None:
                for bi, b in enumerate(blocks):
                    if bi in used_idx:
                        continue
                    if b["start_hms"] == start_hms:
                        chosen = bi
                        break
            if chosen is None:
                continue  # nothing taggable inside this (possibly trimmed) srt
            used_idx.add(chosen)
            tag_for_block[chosen] = (
                f'[ALREADY USED: from {start_hms} to {end_hms or start_hms} '
                f'| phrases used within this frame: "{language_item}"]'
            )
        out = []
        for bi, b in enumerate(blocks):
            if bi in tag_for_block:
                out.append(tag_for_block[bi])
            out.append(b["raw"])
        return "\n\n".join(out) + "\n"
    except Exception:
        try:
            return filter_sub_excluding_approved(srt_content, approved_segments)
        except Exception:
            return srt_content


def _sec_to_srt_ts(sec: float) -> str:
    sec = max(0.0, sec)
    h = int(sec // 3600)
    m = int((sec % 3600) // 60)
    s = sec % 60
    ms = int((s - int(s)) * 1000)
    return f"{h:02d}:{m:02d}:{int(s):02d},{ms:03d}"


def create_clip_subtitle(srt_content: str, start: str, end: str, output_path: str) -> None:
    """
    Extract SRT blocks within [start, end] range, reset timestamps to 0-based, and write to output_path.

    start, end: HH:MM:SS strings
    """
    start_sec = _hms_to_seconds(start)
    end_sec = _hms_to_seconds(end)

    blocks = _parse_srt_blocks(srt_content)
    output_blocks = []

    for block in blocks:
        b_start = block["start_sec"]
        b_end = block["end_sec"]
        if b_end > start_sec and b_start < end_sec:
            output_blocks.append((b_start - start_sec, b_end - start_sec, block["raw"]))

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    ts_pattern = re.compile(
        r"\d{2}:\d{2}:\d{2},\d{3}\s*-->\s*\d{2}:\d{2}:\d{2},\d{3}"
    )
    with open(output_path, "w", encoding="utf-8") as f:
        for idx, (b_start, b_end, raw) in enumerate(output_blocks, start=1):
            lines = raw.strip().splitlines()
            lines[0] = str(idx)
            new_ts = f"{_sec_to_srt_ts(b_start)} --> {_sec_to_srt_ts(b_end)}"
            for i, line in enumerate(lines):
                if ts_pattern.match(line.strip()):
                    lines[i] = new_ts
                    break
            f.write("\n".join(lines) + "\n\n")


def _srt_ts_to_sec(ts: str) -> float:
    """Convert SRT timestamp HH:MM:SS,mmm to float seconds."""
    ts = ts.replace(",", ".")
    parts = ts.strip().split(":")
    return int(parts[0]) * 3600 + int(parts[1]) * 60 + float(parts[2])


def _parse_srt_blocks(srt_content: str) -> list[dict]:
    """Parse SRT into list of {start_hms, end_hms, start_sec, end_sec, raw} dicts."""
    ts_pattern = re.compile(
        r"(\d{2}:\d{2}:\d{2},\d{3})\s*-->\s*(\d{2}:\d{2}:\d{2},\d{3})"
    )
    blocks = []
    raw_blocks = re.split(r"\n\s*\n", srt_content.strip())
    for raw in raw_blocks:
        raw = raw.strip()
        if not raw:
            continue
        match = ts_pattern.search(raw)
        if match:
            s_full = match.group(1)  # HH:MM:SS,mmm
            e_full = match.group(2)
            blocks.append({
                "start_hms": s_full[:8],   # HH:MM:SS — for used_map lookup
                "end_hms":   e_full[:8],
                "start_sec": _srt_ts_to_sec(s_full),
                "end_sec":   _srt_ts_to_sec(e_full),
                "raw": raw,
            })
    return blocks


def expand_sub_context(srt_content: str, timestamp_hms: str, n_lines: int, side: str = "both") -> tuple[str, str]:
    """
    Given a target timestamp (HH:MM:SS) in SRT, find the block containing it,
    then expand ±n_lines dialogue blocks around it.

    side:
      "start" — target is a segment START. On a boundary shared by two contiguous
                blocks, pick the LATER block (the one starting at target) so we
                expand exactly n_lines BEFORE the phrase (not n+1).
      "end"   — target is a segment END. On a boundary, pick the EARLIER block
                (the one ending at target) so we expand exactly n_lines AFTER.
      "both"  — legacy behavior (first containing block).

    Using `<=` on both sides for every call shifts the window earlier on
    boundaries (head too long, tail too short); the side-aware tie-break fixes it.

    Returns (expanded_start_hms, expanded_end_hms).
    If timestamp not found, returns (timestamp_hms, timestamp_hms).
    """
    blocks = _parse_srt_blocks(srt_content)
    if not blocks:
        return timestamp_hms, timestamp_hms

    target_sec = _hms_to_seconds(timestamp_hms)

    # Pick the center block by the matching BOUNDARY, robust to gaps AND overlapping
    # subtitle blocks:
    #   side="start" → block whose START is closest to target (the phrase's first block)
    #   side="end"   → block whose END   is closest to target (the phrase's last block)
    #   side="both"  → block whose START is closest (legacy)
    if side == "end":
        center_idx = min(range(len(blocks)), key=lambda i: abs(blocks[i]["end_sec"] - target_sec))
    else:
        center_idx = min(range(len(blocks)), key=lambda i: abs(blocks[i]["start_sec"] - target_sec))

    start_idx = max(0, center_idx - n_lines)
    end_idx = min(len(blocks) - 1, center_idx + n_lines)

    return blocks[start_idx]["start_hms"], blocks[end_idx]["end_hms"]


# ---------------------------------------------------------------------------
# Timestamp validation
# ---------------------------------------------------------------------------

def validate_timestamp(start: str, end: str, video_duration_sec: float = 0) -> tuple[bool, str]:
    """
    Validate timestamp pair.
    Returns (is_valid, reason_if_invalid).

    Checks:
      1. Format must be HH:MM:SS
      2. start < end
      3. end <= video_duration (only if video_duration_sec > 0)
    """
    if not _is_valid_timestamp(start):
        return False, f"Invalid start format: '{start}' — expected HH:MM:SS"
    if not _is_valid_timestamp(end):
        return False, f"Invalid end format: '{end}' — expected HH:MM:SS"

    start_sec = _hms_to_seconds(start)
    end_sec = _hms_to_seconds(end)

    if start_sec >= end_sec:
        return False, f"start ({start}) must be less than end ({end})"

    if video_duration_sec > 0 and end_sec > video_duration_sec:
        return False, (
            f"end ({end}) exceeds video duration "
            f"({_seconds_to_hms(video_duration_sec)})"
        )

    return True, ""


def normalize_timestamp(val) -> str:
    """
    Normalize any timestamp representation to 'HH:MM:SS'. Returns '' if it cannot
    be parsed (caller decides how to report — never raises).

    Handles the three things openpyxl/Excel can hand back for a time cell:
      - datetime.time / datetime.datetime  → take h/m/s directly
      - a clean string '00:05:32'          → kept
      - a short string '0:5:32'            → padded to '00:05:32'
    """
    import datetime
    if val is None:
        return ""
    if isinstance(val, (datetime.time, datetime.datetime)):
        return f"{val.hour:02d}:{val.minute:02d}:{val.second:02d}"
    parts = str(val).strip().split(":")
    if len(parts) == 3 and all(p.strip().isdigit() for p in parts):
        h, m, s = (int(p) for p in parts)
        return f"{h:02d}:{m:02d}:{s:02d}"
    return ""


def _is_valid_timestamp(ts: str) -> bool:
    """Check if string matches HH:MM:SS format."""
    return bool(re.fullmatch(r"\d{2}:\d{2}:\d{2}", ts.strip()))


def _hms_to_seconds(hms: str) -> float:
    """Convert HH:MM:SS to float seconds."""
    parts = hms.strip().split(":")
    return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])


def _seconds_to_hms(seconds: float) -> str:
    """Convert float seconds to HH:MM:SS."""
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = int(seconds % 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


# ---------------------------------------------------------------------------
# Log
# ---------------------------------------------------------------------------

def save_log(data: list[dict], log_path: str) -> None:
    """Save log data to JSON file."""
    os.makedirs(os.path.dirname(log_path), exist_ok=True)
    with open(log_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_log(log_path: str) -> list[dict]:
    """Load log data from JSON file."""
    with open(log_path, "r", encoding="utf-8") as f:
        return json.load(f)
