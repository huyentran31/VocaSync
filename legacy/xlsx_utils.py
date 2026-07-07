"""
xlsx_utils.py — Excel file utilities for review.xlsx.

Functions:
  export_xlsx          — Create/overwrite review.xlsx with clip data
  update_xlsx_links    — Update link_video, link_sub, sub_lines after B3 cut
  read_xlsx            — Read review.xlsx with type casting and validation
  validate_xlsx        — Validate teacher edits before pipeline runs
  detect_file_lock     — Check if Excel has review.xlsx open (~$review.xlsx)
  apply_highlight      — Apply row highlighting (blue/red/white)
  get_global_max_attempt — Max attempt number across entire xlsx
  get_global_max_index   — Max global_index across entire xlsx
"""

import os
import re
import sys
from datetime import datetime
import pathlib

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
import config

# Column order in xlsx
COLUMNS = [
    "video", "clip", "attempt", "status",
    "start", "end",
    "system_note",
    "language_item", "original_word", "language_tag",
    "ai_note", "example_sub",
    "comment", "sub_lines", "new_start", "new_end", "merge_order",
    "link_video", "link_sub",
    "global_index",
]

FREEZE_COLS = 4  # freeze video, clip, attempt, status

STATUS_VALUES = {"needs_revision", "manual_edit", "pending_review", "approved", "rejected", "fail"}

FILL_BLUE = PatternFill("solid", fgColor="DBEAFE")
FILL_RED = PatternFill("solid", fgColor="FEE2E2")
FILL_ORANGE = PatternFill("solid", fgColor="FEF3C7")
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")


def detect_file_lock(xlsx_path: str) -> bool:
    """
    Check if Excel has the file open by looking for ~$review.xlsx.
    Returns True if file is locked.
    """
    dir_path = os.path.dirname(xlsx_path)
    lock_file = os.path.join(dir_path, "~$review.xlsx")
    return os.path.exists(lock_file)


def export_xlsx(clips: list[dict], xlsx_path: str, pre_cut: bool = True) -> None:
    """
    Create review.xlsx from list of clip dicts.

    pre_cut=True  → link_video, link_sub are empty (B2.5 Step 2 export)
    pre_cut=False → link_video, link_sub filled from clip data

    clips: list of dicts with keys matching COLUMNS.
    """
    if detect_file_lock(xlsx_path):
        raise RuntimeError("review.xlsx is open in Excel. Please close it and try again.")

    _dir = os.path.dirname(os.path.abspath(xlsx_path))
    os.makedirs(_dir, exist_ok=True)

    rows = []
    for clip in clips:
        row = {col: clip.get(col, "") for col in COLUMNS}
        if pre_cut:
            row["link_video"] = ""
            row["link_sub"] = ""
        else:
            # Build hyperlinks
            clip_path = clip.get("clip_path", "")
            srt_path = clip.get("clip_srt_path", "")
            row["link_video"] = clip_path
            row["link_sub"] = srt_path
        rows.append(row)

    df = pd.DataFrame(rows, columns=COLUMNS)

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Review")
        wb = writer.book
        ws = writer.sheets["Review"]

        # Freeze first 4 columns + header row
        ws.freeze_panes = ws.cell(row=2, column=FREEZE_COLS + 1)

        # Format hyperlink cells
        link_col_video = COLUMNS.index("link_video") + 1
        link_col_sub = COLUMNS.index("link_sub") + 1

        for row_idx, clip in enumerate(clips, start=2):
            clip_path = clip.get("clip_path", "")
            srt_path = clip.get("clip_srt_path", "")

            if not pre_cut and clip_path and os.path.exists(clip_path):
                cell = ws.cell(row=row_idx, column=link_col_video)
                abs_path = os.path.abspath(clip_path)
                cell.hyperlink = pathlib.Path(abs_path).as_uri()
                cell.value = os.path.basename(clip_path)
                cell.font = Font(color="0563C1", underline="single")

            if not pre_cut and srt_path and os.path.exists(srt_path):
                cell = ws.cell(row=row_idx, column=link_col_sub)
                abs_path = os.path.abspath(srt_path)
                cell.hyperlink = pathlib.Path(abs_path).as_uri()
                cell.value = os.path.basename(srt_path)
                cell.font = Font(color="0563C1", underline="single")

        # Format new_start / new_end as Text to prevent Excel auto-conversion
        new_start_col = COLUMNS.index("new_start") + 1
        new_end_col = COLUMNS.index("new_end") + 1
        for row_idx in range(2, len(clips) + 2):
            ws.cell(row=row_idx, column=new_start_col).number_format = "@"
            ws.cell(row=row_idx, column=new_end_col).number_format = "@"

        # Add _instructions sheet
        ws_inst = wb.create_sheet("_instructions")
        instructions = [
            ["AI Teaching System — Flow 1 | review.xlsx Instructions"],
            [""],
            ["STATUS VALUES"],
            ["needs_revision", "Send back to AI for a different segment"],
            ["manual_edit", "Enter new timestamps manually in new_start/new_end columns"],
            ["pending_review", "Processed — awaiting your review (highlighted blue)"],
            ["approved", "Include in final video"],
            ["rejected", "Exclude from final (set by system at Finalize)"],
            ["fail", "Processing error — see system_note for details"],
            [""],
            ["RULES"],
            ["• Only ONE approved row per clip (identified by clip column)"],
            ["• manual_edit requires BOTH new_start AND new_end — format HH:MM:SS"],
            ["• sub_lines must be a positive integer"],
            ["• merge_order must be a positive integer if filled"],
            ["• Do NOT manually set status = fail — use needs_revision or manual_edit instead"],
            ["• comment column: your feedback only — never edited by system"],
            ["• system_note: system error messages only — do not edit"],
            [""],
            ["MERGE ORDER"],
            ["• Fill merge_order to control clip order in final video"],
            ["• Leave blank to use default sort (group → video → timestamp)"],
            ["• Duplicate merge_order values will block Finalize"],
            [""],
            ["LINK_VIDEO (Mac users)"],
            ["• Right-click the link → Edit Hyperlink → copy path"],
        ]
        for row in instructions:
            ws_inst.append(row)

        # Column widths
        col_widths = {
            "video": 25, "clip": 35, "attempt": 8, "status": 15,
            "start": 10, "end": 10, "system_note": 40, "ai_note": 40,
            "example_sub": 50, "language_item": 20, "language_tag": 20,
            "original_word": 20, "comment": 40, "sub_lines": 10,
            "new_start": 12, "new_end": 12, "merge_order": 12,
            "link_video": 30, "link_sub": 30, "global_index": 12,
        }
        for col_name, width in col_widths.items():
            if col_name in COLUMNS:
                col_letter = get_column_letter(COLUMNS.index(col_name) + 1)
                ws.column_dimensions[col_letter].width = width

    _apply_initial_highlight(xlsx_path)


def update_xlsx_links(xlsx_path: str, clip_updates: list[dict]) -> None:
    """
    Update xlsx with clip, global_index, attempt, status, link_video, link_sub, sub_lines.

    clip_updates: list of {
        video, start, end,            <- match key (composite)
        clip,                         <- clip filename
        global_index,                 <- assigned index
        attempt,                      <- attempt number
        status,                       <- e.g. pending_review / fail
        clip_path: abs_path,          <- for hyperlink
        clip_srt_path: abs_path,      <- for hyperlink
        sub_lines: int
    }
    Matches rows by (video, start, end).
    """
    if detect_file_lock(xlsx_path):
        raise RuntimeError("review.xlsx is open in Excel. Please close it and try again.")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Review"]

    # Build header map
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1) if cell.value}

    video_col = headers.get("video")
    start_col = headers.get("start")
    end_col = headers.get("end")
    clip_col = headers.get("clip")
    global_index_col = headers.get("global_index")
    attempt_col = headers.get("attempt")
    status_col = headers.get("status")
    link_video_col = headers.get("link_video")
    link_sub_col = headers.get("link_sub")
    sub_lines_col = headers.get("sub_lines")

    # Build update lookup by (video, start, end) composite key
    update_map = {
        (str(u.get("video", "")).strip(),
         str(u.get("start", "")).strip(),
         str(u.get("end", "")).strip()): u
        for u in clip_updates
    }

    for row in ws.iter_rows(min_row=2):
        row_video = str(row[video_col - 1].value or "").strip() if video_col else ""
        row_start = str(row[start_col - 1].value or "").strip() if start_col else ""
        row_end = str(row[end_col - 1].value or "").strip() if end_col else ""
        key = (row_video, row_start, row_end)

        if key not in update_map:
            continue

        upd = update_map[key]

        if clip_col and upd.get("clip"):
            row[clip_col - 1].value = upd["clip"]
        if global_index_col and upd.get("global_index") is not None:
            row[global_index_col - 1].value = upd["global_index"]
        if attempt_col and upd.get("attempt") is not None:
            row[attempt_col - 1].value = upd["attempt"]
        if status_col and upd.get("status"):
            row[status_col - 1].value = upd["status"]
        if sub_lines_col and upd.get("sub_lines") is not None:
            row[sub_lines_col - 1].value = upd["sub_lines"]

        if link_video_col and upd.get("clip_path") and os.path.exists(upd["clip_path"]):
            cell = row[link_video_col - 1]
            abs_path = os.path.abspath(upd["clip_path"])
            cell.hyperlink = pathlib.Path(abs_path).as_uri()
            cell.value = os.path.basename(upd["clip_path"])
            cell.font = Font(color="0563C1", underline="single")

        if link_sub_col and upd.get("clip_srt_path") and os.path.exists(upd["clip_srt_path"]):
            cell = row[link_sub_col - 1]
            abs_path = os.path.abspath(upd["clip_srt_path"])
            cell.hyperlink = pathlib.Path(abs_path).as_uri()
            cell.value = os.path.basename(upd["clip_srt_path"])
            cell.font = Font(color="0563C1", underline="single")

    wb.save(xlsx_path)


def update_xlsx_hyperlinks_by_clip(xlsx_path: str, clip_updates: list[dict]) -> None:
    """
    After B4 append_rows_to_xlsx, update hyperlinks for newly appended rows.
    Matches rows by clip column value (safe to use here because clip is already written
    by append_rows_to_xlsx before this is called).

    clip_updates: list of {clip: filename, clip_path: abs_path, clip_srt_path: abs_path}
    """
    if detect_file_lock(xlsx_path):
        raise RuntimeError("review.xlsx is open in Excel. Please close it and try again.")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Review"]

    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1) if cell.value}
    clip_col = headers.get("clip")
    link_video_col = headers.get("link_video")
    link_sub_col = headers.get("link_sub")

    update_map = {u["clip"]: u for u in clip_updates}

    for row in ws.iter_rows(min_row=2):
        clip_val = row[clip_col - 1].value if clip_col else None
        if not clip_val or str(clip_val).strip() not in update_map:
            continue

        upd = update_map[str(clip_val).strip()]

        if link_video_col and upd.get("clip_path") and os.path.exists(upd["clip_path"]):
            cell = row[link_video_col - 1]
            abs_path = os.path.abspath(upd["clip_path"])
            cell.hyperlink = pathlib.Path(abs_path).as_uri()
            cell.value = os.path.basename(upd["clip_path"])
            cell.font = Font(color="0563C1", underline="single")

        if link_sub_col and upd.get("clip_srt_path") and os.path.exists(upd["clip_srt_path"]):
            cell = row[link_sub_col - 1]
            abs_path = os.path.abspath(upd["clip_srt_path"])
            cell.hyperlink = pathlib.Path(abs_path).as_uri()
            cell.value = os.path.basename(upd["clip_srt_path"])
            cell.font = Font(color="0563C1", underline="single")

    wb.save(xlsx_path)


def read_xlsx(xlsx_path: str) -> pd.DataFrame:
    """
    Read review.xlsx with type casting.
    Returns DataFrame with all COLUMNS present.
    Retries on BadZipFile — happens when the file is read while still being
    written (mid-save race), not real corruption.
    """
    import time as _time
    last_err = None
    for _attempt in range(5):
        try:
            df = pd.read_excel(xlsx_path, sheet_name="Review", engine="openpyxl", dtype=str)
            break
        except Exception as e:
            last_err = e
            _time.sleep(0.4)
    else:
        raise last_err
    df = df.fillna("")

    # Type cast numeric columns
    for col in ["attempt", "sub_lines", "merge_order", "global_index"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    return df


def validate_xlsx(xlsx_path: str) -> list[dict]:
    """
    Validate teacher edits. Returns list of error dicts:
      {"row": int, "col": str, "message": str}

    Rules:
      - manual_edit: both new_start AND new_end must be filled
      - new_start / new_end: if filled, must be HH:MM:SS
      - sub_lines: must be positive integer if filled
      - merge_order: must be positive integer if filled
      - attempt: must be positive integer
    """
    if detect_file_lock(xlsx_path):
        raise RuntimeError("review.xlsx is open in Excel. Please close it and try again.")

    df = read_xlsx(xlsx_path)
    errors = []

    ts_pattern = re.compile(r"^\d{2}:\d{2}:\d{2}$")

    for idx, row in df.iterrows():
        row_num = idx + 2  # Excel row (1-indexed + header)

        status = str(row.get("status", "")).strip()

        # manual_edit: both new_start AND new_end required
        if status == "manual_edit":
            new_start = str(row.get("new_start", "")).strip()
            new_end = str(row.get("new_end", "")).strip()
            if not new_start or not new_end:
                errors.append({"row": row_num, "col": "new_start", "message": "manual_edit requires both new_start and new_end"})
                errors.append({"row": row_num, "col": "new_end", "message": "manual_edit requires both new_start and new_end"})
            else:
                if not ts_pattern.match(new_start):
                    errors.append({"row": row_num, "col": "new_start", "message": f"Invalid format '{new_start}' — expected HH:MM:SS"})
                if not ts_pattern.match(new_end):
                    errors.append({"row": row_num, "col": "new_end", "message": f"Invalid format '{new_end}' — expected HH:MM:SS"})
                if ts_pattern.match(new_start) and ts_pattern.match(new_end):
                    def _hms_to_sec(t):
                        p = t.split(":")
                        return int(p[0]) * 3600 + int(p[1]) * 60 + int(p[2])
                    if _hms_to_sec(new_start) >= _hms_to_sec(new_end):
                        errors.append({"row": row_num, "col": "new_start", "message": f"new_start ({new_start}) must be less than new_end ({new_end})"})

        # new_start / new_end format if filled (non-manual_edit)
        if status != "manual_edit":
            for col_name in ["new_start", "new_end"]:
                val = str(row.get(col_name, "")).strip()
                if val and not ts_pattern.match(val):
                    errors.append({"row": row_num, "col": col_name, "message": f"Invalid format '{val}' — expected HH:MM:SS"})

        # sub_lines
        sub_lines = row.get("sub_lines")
        if pd.notna(sub_lines) and str(sub_lines).strip():
            try:
                v = int(float(str(sub_lines)))
                if v <= 0:
                    raise ValueError
            except (ValueError, TypeError):
                errors.append({"row": row_num, "col": "sub_lines", "message": f"sub_lines must be a positive integer, got '{sub_lines}'"})

        # merge_order
        merge_order = row.get("merge_order")
        if pd.notna(merge_order) and str(merge_order).strip():
            try:
                v = int(float(str(merge_order)))
                if v <= 0:
                    raise ValueError
            except (ValueError, TypeError):
                errors.append({"row": row_num, "col": "merge_order", "message": f"merge_order must be a positive integer, got '{merge_order}'"})

        # attempt
        attempt = row.get("attempt")
        if pd.notna(attempt) and str(attempt).strip():
            try:
                v = int(float(str(attempt)))
                if v <= 0:
                    raise ValueError
            except (ValueError, TypeError):
                errors.append({"row": row_num, "col": "attempt", "message": f"attempt must be a positive integer, got '{attempt}'"})

    return errors


def apply_highlight(xlsx_path: str) -> None:
    """
    Apply row highlighting based on status and attempt:
      Blue  (#DBEAFE): status == pending_review AND attempt == global_max
      Red   (#FEE2E2): status == fail AND attempt == global_max
      White: all others
    """
    if detect_file_lock(xlsx_path):
        raise RuntimeError("review.xlsx is open in Excel. Please close it and try again.")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Review"]

    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1) if cell.value}
    status_col = headers.get("status")
    attempt_col = headers.get("attempt")
    max_col = ws.max_column

    # Pass 1: find global max attempt
    global_max = 0
    for row in ws.iter_rows(min_row=2):
        if attempt_col:
            val = row[attempt_col - 1].value
            try:
                v = int(val)
                if v > global_max:
                    global_max = v
            except (TypeError, ValueError):
                pass

    # Pass 2: find duplicate original_words (2nd+ occurrence in row order)
    original_word_col = headers.get("original_word")
    duplicate_row_nums = set()  # excel row numbers to tint orange
    if original_word_col:
        first_seen: dict[str, int] = {}
        for row in ws.iter_rows(min_row=2):
            word_val = row[original_word_col - 1].value
            if word_val:
                word = str(word_val).strip().lower()
                if word:
                    if word in first_seen:
                        duplicate_row_nums.add(row[0].row)
                    else:
                        first_seen[word] = row[0].row

    # Pass 3: apply colors
    ACTIVE_STATUSES = {"pending_review", "needs_revision", "manual_edit"}
    for row in ws.iter_rows(min_row=2):
        status_val = row[status_col - 1].value if status_col else ""
        attempt_val = None
        if attempt_col:
            try:
                attempt_val = int(row[attempt_col - 1].value)
            except (TypeError, ValueError):
                pass

        is_current = (attempt_val == global_max)

        if status_val == "pending_review" and is_current:
            fill = FILL_BLUE
        elif status_val == "fail" and is_current:
            fill = FILL_RED
        elif status_val in ACTIVE_STATUSES and row[0].row in duplicate_row_nums:
            fill = FILL_ORANGE
        else:
            fill = FILL_WHITE

        for cell in row[:max_col]:
            cell.fill = fill

    wb.save(xlsx_path)


def _apply_initial_highlight(xlsx_path: str) -> None:
    """Apply initial highlighting after export_xlsx."""
    apply_highlight(xlsx_path)


def highlight_error_cells(xlsx_path: str, errors: list[dict]) -> None:
    """Highlight specific cells red for validation errors."""
    if not errors:
        return
    if detect_file_lock(xlsx_path):
        return

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Review"]
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1) if cell.value}

    for error in errors:
        row_num = error["row"]
        col_name = error["col"]
        col_idx = headers.get(col_name)
        if col_idx:
            ws.cell(row=row_num, column=col_idx).fill = FILL_RED

    wb.save(xlsx_path)


def get_global_max_attempt(xlsx_path: str) -> int:
    """Return max attempt number across entire xlsx."""
    df = read_xlsx(xlsx_path)
    if "attempt" not in df.columns or df["attempt"].dropna().empty:
        return 0
    return int(df["attempt"].dropna().max())


def get_global_max_index(xlsx_path: str) -> int:
    """Return max global_index across entire xlsx."""
    df = read_xlsx(xlsx_path)
    if "global_index" not in df.columns or df["global_index"].dropna().empty:
        return 0
    return int(df["global_index"].dropna().max())


def update_status_in_xlsx(xlsx_path: str, from_status: str, to_status: str) -> int:
    """
    Update all rows with status == from_status to to_status in review.xlsx.
    Preserves all formatting, hyperlinks, and other cell values.
    Returns number of rows changed.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Review"]

    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1) if cell.value}
    status_col = headers.get("status")
    if status_col is None:
        return 0

    count = 0
    for row in ws.iter_rows(min_row=2):
        cell = row[status_col - 1]
        if cell.value == from_status:
            cell.value = to_status
            count += 1

    wb.save(xlsx_path)
    return count


def set_status_by_rownums(xlsx_path: str, rownums: list[int], to_status: str) -> int:
    """
    Set status = to_status for specific Excel row numbers (1-based, including header).
    Used to mark successfully-processed manual_edit source rows as 'rejected' so they
    are kept as a record but not re-cut on the next Process Next Round.
    Returns number of rows changed.
    """
    if not rownums:
        return 0
    if detect_file_lock(xlsx_path):
        raise RuntimeError("review.xlsx is open in Excel. Please close it and try again.")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Review"]

    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1) if cell.value}
    status_col = headers.get("status")
    if status_col is None:
        return 0

    count = 0
    for rn in set(rownums):
        ws.cell(row=rn, column=status_col).value = to_status
        count += 1

    wb.save(xlsx_path)
    return count


def set_cell_values(xlsx_path: str, col_name: str, value_by_rownum: dict) -> int:
    """
    Set a single column's value for specific Excel row numbers (1-based, incl header).
    value_by_rownum: {rownum: value}. Returns number of cells changed.
    Used by the language_item/duplicate re-check pass to refresh system_note.
    """
    if not value_by_rownum:
        return 0
    if detect_file_lock(xlsx_path):
        raise RuntimeError("review.xlsx is open in Excel. Please close it and try again.")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Review"]

    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1) if cell.value}
    col = headers.get(col_name)
    if col is None:
        return 0

    count = 0
    for rn, val in value_by_rownum.items():
        ws.cell(row=rn, column=col).value = val
        count += 1

    wb.save(xlsx_path)
    return count


def append_rows_to_xlsx(xlsx_path: str, new_rows: list[dict]) -> None:
    """
    Append new rows to existing review.xlsx (for B4 next attempts and B5* Find More).
    new_rows: list of dicts matching COLUMNS.
    """
    if detect_file_lock(xlsx_path):
        raise RuntimeError("review.xlsx is open in Excel. Please close it and try again.")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Review"]

    headers = [cell.value for cell in ws[1]]

    for row_dict in new_rows:
        row_values = [row_dict.get(col, "") for col in headers]
        ws.append(row_values)

    wb.save(xlsx_path)
    apply_highlight(xlsx_path)
